import requests
from bs4 import BeautifulSoup
import openpyxl

COMPANY_CODES = [
    '036570',
    '194480',
    '263750',
    '112040',
    '293490'
]

wb = openpyxl.Workbook()
ws = wb.active

ws['A1'] = '회사 이름'
ws['B1'] = '현재가'

for idx in range(len(COMPANY_CODES)):

    response = requests.get(f'https://finance.naver.com/item/sise.naver?code={COMPANY_CODES[idx]}')
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')

    company_name = soup.select_one('#middle > div.h_company > div.wrap_company > h2 > a')
    price = soup.select_one('strong#_nowVal')
    print(company_name.text)
    print(price.text)
    ws[f'A{idx+2}'] = company_name.text
    ws[f'B{idx+2}'] = price.text

wb.save('현재 주가.xlsx')